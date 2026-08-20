from __future__ import annotations
import os
import uuid
import pytest
from pathlib import Path
from fastapi.testclient import TestClient
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker, Session
from sqlalchemy.pool import StaticPool
import openpyxl
import io
from app.main import app
from app.models import Base
from app.models.user import User
from app.models.course import Course
from app.models.material import Material
from app.models.question import Question, Option
from app.models.system import Export
from app.services.export_service import _generate_excel_workbook, _sanitize_for_excel
from app.core.storage import get_export_dir
from tests.test_question_bank_api import _make_user

client = TestClient(app)

@pytest.fixture()
def db_engine():
    engine = create_engine(
        "sqlite://",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(bind=engine)
    return engine

@pytest.fixture()
def db_session_factory(db_engine):
    return sessionmaker(autocommit=False, autoflush=False, bind=db_engine)

@pytest.fixture()
def db(db_session_factory):
    session = db_session_factory()
    try:
        yield session
    finally:
        session.close()

@pytest.fixture(autouse=True)
def override_deps(db_session_factory):
    from app.api.deps import get_db
    def override_db():
        session = db_session_factory()
        try:
            yield session
        finally:
            session.close()
    app.dependency_overrides[get_db] = override_db
    yield
    app.dependency_overrides.clear()

@pytest.fixture(autouse=True)
def mock_export_dir(monkeypatch, tmp_path):
    # Dùng tmp_path làm EXPORT_DIR cho toàn bộ suite test export
    monkeypatch.setenv("EXPORT_DIR", str(tmp_path / "exports"))
    yield tmp_path / "exports"

# Helper fixtures
@pytest.fixture
def lecturer_user(db):
    user = _make_user(db, email="lecturer_export@test.com", role="lecturer")
    return user

@pytest.fixture
def other_lecturer_user(db):
    return _make_user(db, email="other_lecturer_export@test.com", role="lecturer")

@pytest.fixture
def test_course(db, lecturer_user):
    c = Course(title="Export Course", created_by=lecturer_user.id)
    db.add(c)
    db.commit()
    db.refresh(c)
    return c

@pytest.fixture
def test_material(db, test_course, lecturer_user):
    m = Material(
        title="Test Material Source",
        file_path="/fake/path",
        course_id=test_course.id,
        uploaded_by=lecturer_user.id
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return m

def _create_question(db, course_id, material_id, content, status="approved", diff="medium", num_opts=4, correct_idx=0):
    q = Question(
        course_id=course_id,
        material_id=material_id,
        content=content,
        status=status,
        difficulty=diff,
        bloom_level="remember",
        question_type="multiple_choice",
        explanation=f"Giải thích cho {content}"
    )
    db.add(q)
    db.commit()
    db.refresh(q)
    
    # Adding options
    for i in range(num_opts):
        opt = Option(
            question_id=q.id,
            content=f"Option {i} for {content}",
            is_correct=(i == correct_idx)
        )
        db.add(opt)
    db.commit()
    return q

# Unit tests for _sanitize_for_excel
def test_sanitize_for_excel():
    assert _sanitize_for_excel(None) == ""
    assert _sanitize_for_excel("") == ""
    assert _sanitize_for_excel("Hello") == "Hello"
    assert _sanitize_for_excel("1+1") == "1+1"
    assert _sanitize_for_excel("=SUM(A1:A2)") == "'=SUM(A1:A2)"
    assert _sanitize_for_excel("+123") == "'+123"
    assert _sanitize_for_excel("-456") == "'-456"
    assert _sanitize_for_excel("@something") == "'@something"

# Unit tests for workbook generation
def test_generate_excel_workbook_empty():
    wb = _generate_excel_workbook([])
    assert "Câu hỏi" in wb.sheetnames
    ws = wb["Câu hỏi"]
    assert ws.max_row == 1
    headers = [cell.value for cell in ws[1]]
    assert "Đáp án D" in headers
    assert "Đáp án E" not in headers

def test_generate_excel_workbook_with_data(db, test_course, test_material):
    q1 = _create_question(db, test_course.id, test_material.id, "Lịch sử Tiếng Việt") 
    q2 = _create_question(db, test_course.id, test_material.id, "=Formula injection", num_opts=2) 
    q3 = _create_question(db, test_course.id, test_material.id, "Extra opts", num_opts=5, correct_idx=4)
    
    db.refresh(q1)
    db.refresh(q2)
    db.refresh(q3)
    
    questions = [q1, q2, q3]
    
    wb = _generate_excel_workbook(questions)
    ws = wb["Câu hỏi"]
    
    assert ws.max_row == 4
    headers = [cell.value for cell in ws[1]]
    assert "Đáp án E" in headers
    
    assert ws.cell(row=2, column=2).value == "Lịch sử Tiếng Việt"
    assert ws.cell(row=2, column=8).value == "A"
    assert ws.cell(row=2, column=13).value == test_material.title
    
    assert ws.cell(row=3, column=2).value == "'=Formula injection"
    assert ws.cell(row=3, column=3).value == "Option 0 for =Formula injection"
    assert ws.cell(row=3, column=4).value == "Option 1 for =Formula injection"
    assert ws.cell(row=3, column=5).value is None or ws.cell(row=3, column=5).value == ""
    assert ws.cell(row=3, column=8).value == "A"
    
    assert ws.cell(row=4, column=7).value == "Option 4 for Extra opts"
    assert ws.cell(row=4, column=8).value == "E"

# API Integration tests
def get_auth_headers(user: User):
    from app.core.security import create_access_token
    token = create_access_token(user.email)
    return {"Authorization": f"Bearer {token}"}

def test_api_export_excel_success(db, lecturer_user, test_course, test_material, mock_export_dir):
    q1 = _create_question(db, test_course.id, test_material.id, "Q1")
    q2 = _create_question(db, test_course.id, test_material.id, "Q2")
    q3 = _create_question(db, test_course.id, test_material.id, "Q3")
    
    headers = get_auth_headers(lecturer_user)
    
    request_data = {"question_ids": [q3.id, q1.id, q3.id, q2.id]}
    
    response = client.post("/api/v1/exports/excel", json=request_data, headers=headers)
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "questions_export" in response.headers["content-disposition"]
    
    stream = io.BytesIO(response.content)
    wb = openpyxl.load_workbook(stream)
    ws = wb.active
    assert ws.max_row == 4
    assert ws.cell(row=2, column=2).value == "Q3"
    assert ws.cell(row=3, column=2).value == "Q1"
    assert ws.cell(row=4, column=2).value == "Q2"
    
    # Check DB record
    export_record = db.query(Export).first()
    assert export_record is not None
    assert export_record.course_id == test_course.id
    assert export_record.exported_by == lecturer_user.id
    assert export_record.format == "xlsx"
    assert export_record.question_ids == [q3.id, q1.id, q2.id] 
    
    # Check file exists
    assert (mock_export_dir / export_record.file_path).exists()
    assert "/" not in export_record.file_path
    assert "\\" not in export_record.file_path

def test_api_export_excel_multi_course(db, lecturer_user, test_course, test_material, mock_export_dir):
    c2 = Course(title="Course 2", created_by=lecturer_user.id)
    db.add(c2)
    db.commit()
    db.refresh(c2)
    
    q1 = _create_question(db, test_course.id, test_material.id, "Q1")
    q2 = _create_question(db, c2.id, test_material.id, "Q2")
    
    headers = get_auth_headers(lecturer_user)
    response = client.post("/api/v1/exports/excel", json={"question_ids": [q1.id, q2.id]}, headers=headers)
    assert response.status_code == 200
    
    export_record = db.query(Export).first()
    assert export_record.course_id is None # Mixed courses -> NULL

def test_api_export_excel_rollback_on_db_fail(db, lecturer_user, test_course, test_material, monkeypatch, mock_export_dir):
    q1 = _create_question(db, test_course.id, test_material.id, "Q1")
    headers = get_auth_headers(lecturer_user)
    
    original_add = Session.add
    def mock_add(*args, **kwargs):
        raise Exception("Mock DB Failure")
        
    monkeypatch.setattr(Session, "add", mock_add)
    
    response = client.post("/api/v1/exports/excel", json={"question_ids": [q1.id]}, headers=headers)
    assert response.status_code == 500
    
    # Temp and final file should be removed
    files = list(mock_export_dir.iterdir()) if mock_export_dir.exists() else []
    assert len(files) == 0

def test_history_api(db, lecturer_user, other_lecturer_user, test_course):
    # Add some records
    e1 = Export(course_id=test_course.id, exported_by=lecturer_user.id, file_path="f1.xlsx", question_ids=[1])
    e2 = Export(course_id=None, exported_by=lecturer_user.id, file_path="f2.xlsx", question_ids=[2,3])
    e3 = Export(course_id=test_course.id, exported_by=other_lecturer_user.id, file_path="f3.xlsx", question_ids=[4])
    db.add_all([e1, e2, e3])
    db.commit()
    
    headers = get_auth_headers(lecturer_user)
    res = client.get("/api/v1/exports?skip=0&limit=10", headers=headers)
    assert res.status_code == 200
    data = res.json()
    assert data["total"] == 2
    assert len(data["items"]) == 2
    
    assert data["items"][0]["file_name"] in ["f1.xlsx", "f2.xlsx"]
    assert "file_path" not in data["items"][0]
    
def test_history_api_validation(lecturer_user):
    headers = get_auth_headers(lecturer_user)
    assert client.get("/api/v1/exports?skip=-1", headers=headers).status_code == 422
    assert client.get("/api/v1/exports?limit=0", headers=headers).status_code == 422
    assert client.get("/api/v1/exports?limit=101", headers=headers).status_code == 422

def test_download_api(db, lecturer_user, other_lecturer_user, test_course, mock_export_dir):
    filename = "test_download.xlsx"
    filepath = mock_export_dir / filename
    mock_export_dir.mkdir(exist_ok=True)
    filepath.write_text("dummy excel")
    
    e1 = Export(course_id=test_course.id, exported_by=lecturer_user.id, file_path=filename, question_ids=[1])
    db.add(e1)
    db.commit()
    db.refresh(e1)
    
    headers = get_auth_headers(lecturer_user)
    res = client.get(f"/api/v1/exports/{e1.id}/download", headers=headers)
    assert res.status_code == 200
    assert res.content == b"dummy excel"
    assert "attachment" in res.headers["content-disposition"]
    
    # User not owner
    other_headers = get_auth_headers(other_lecturer_user)
    res = client.get(f"/api/v1/exports/{e1.id}/download", headers=other_headers)
    assert res.status_code == 404
    
    # ID not found
    res = client.get(f"/api/v1/exports/9999/download", headers=headers)
    assert res.status_code == 404
    
    # File missing physically
    filepath.unlink()
    res = client.get(f"/api/v1/exports/{e1.id}/download", headers=headers)
    assert res.status_code == 404

def test_download_api_path_traversal(db, lecturer_user, test_course, mock_export_dir):
    e1 = Export(course_id=test_course.id, exported_by=lecturer_user.id, file_path="../../test_download.xlsx", question_ids=[1])
    db.add(e1)
    db.commit()
    db.refresh(e1)
    
    headers = get_auth_headers(lecturer_user)
    res = client.get(f"/api/v1/exports/{e1.id}/download", headers=headers)
    assert res.status_code == 404
